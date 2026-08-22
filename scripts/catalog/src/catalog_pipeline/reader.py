from __future__ import annotations

import csv
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from catalog_pipeline.constants import REQUIRED_COLUMNS
from catalog_pipeline.errors import CatalogReadError
from catalog_pipeline.models import SourceRow


def format_price(value: Decimal) -> str:
    return f"{value.quantize(Decimal('0.01'))}"


def parse_price(raw: object) -> Decimal:
    if raw is None or raw == "":
        raise CatalogReadError("Price is required.")
    try:
        return Decimal(str(raw).strip().replace(",", ".")).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError) as exc:
        raise CatalogReadError(f"Invalid price: {raw!r}") from exc


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _require_headers(headers: list[str]) -> dict[str, int]:
    index = {name.strip().lower(): i for i, name in enumerate(headers) if name and name.strip()}
    missing = [col for col in REQUIRED_COLUMNS if col.lower() not in index]
    if missing:
        raise CatalogReadError(f"Missing required columns: {', '.join(missing)}")
    return {col: index[col.lower()] for col in REQUIRED_COLUMNS}


def _row_is_empty(values: list[object]) -> bool:
    return all(_cell_text(v) == "" for v in values)


def _rows_to_source(raw_rows: list[tuple[int, list[object]]], column_index: dict[str, int]) -> list[SourceRow]:
    seen: dict[str, int] = {}
    products: list[SourceRow] = []
    for row_number, values in raw_rows:
        if _row_is_empty(values):
            continue
        sku = _cell_text(_at(values, column_index["SKU"]))
        name = _cell_text(_at(values, column_index["Name"]))
        description = _cell_text(_at(values, column_index["Description"]))
        collection = _cell_text(_at(values, column_index["Collection"]))
        price_raw = _at(values, column_index["Price"])
        if not sku:
            raise CatalogReadError(f"Row {row_number}: SKU is required.")
        sku_key = sku.casefold()
        if sku_key in seen:
            raise CatalogReadError(f"Row {row_number}: duplicate SKU {sku!r} (also row {seen[sku_key]}).")
        seen[sku_key] = row_number
        products.append(
            SourceRow(
                sku=sku,
                name=name,
                description=description,
                price=parse_price(price_raw),
                collection_name=collection,
            )
        )
    return products


def _at(values: list[object], index: int) -> object:
    if index >= len(values):
        return None
    return values[index]


def read_xlsx(path: Path) -> list[SourceRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise CatalogReadError(f"{path}: empty workbook.") from exc
        headers = [_cell_text(v) for v in header_row]
        column_index = _require_headers(headers)
        raw_rows: list[tuple[int, list[object]]] = []
        for offset, row in enumerate(rows_iter, start=2):
            raw_rows.append((offset, list(row)))
    finally:
        workbook.close()
    return _rows_to_source(raw_rows, column_index)


def read_csv(path: Path) -> list[SourceRow]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            header_row = next(reader)
        except StopIteration as exc:
            raise CatalogReadError(f"{path}: empty CSV.") from exc
        column_index = _require_headers([_cell_text(v) for v in header_row])
        raw_rows = [(i, list(row)) for i, row in enumerate(reader, start=2)]
    return _rows_to_source(raw_rows, column_index)


def read_export(path: Path | str) -> list[SourceRow]:
    source = Path(path)
    if not source.exists():
        raise CatalogReadError(f"Export not found: {source}")
    suffix = source.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(source)
    if suffix == ".csv":
        return read_csv(source)
    raise CatalogReadError(f"Unsupported export type: {source.suffix}")
